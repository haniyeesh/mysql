import openpyxl
import datetime
import mysql.connector as sql


def back_up_teacher():
    w = openpyxl.Workbook()
    s = w.active
    s.title = "teacher"
    s['a1'] = 'name'
    s['b1'] = 'last name'
    s['c1'] = 'code'
    s['d1'] = 'bday '
    s['e1'] = 'phone num'
    s['f1'] = 'madrak'
    s['g1'] = 'rooz class'
    s['h1'] = 'lesson'
    s['i1'] = 'bimeh'

    try:                                                    #etesal be databse(mysql)
        con = sql.connect(host='localhost', user='root', password='P@ssw0rd', database="amozeshgah")
        cur = con.cursor()
        cur.execute('SELECT * FROM teacher')
        lst_code = cur.fetchall()
        for i in range(0, len(lst_code)):       #har tuple
            for j in range(0, len(lst_code[i])):        #har item tuple
                s.cell(row=i+2, column=j+1).value = lst_code[i][j]
        w.save('c:\\haniye\\jalase10.xlsx')
    except Exception as x:
        print(x)


def back_up_student():
    w = openpyxl.Workbook()
    s = w.active
    s.title = "student"
    s['a1'] = 'name'
    s['b1'] = 'last name'
    s['c1'] = 'code'
    s['d1'] = 'bday '
    s['e1'] = 'phone num'
    try:                                                    #etesal be databse(mysql)

        con = sql.connect(host='localhost', user='root', password='P@ssw0rd', database="amozeshgah")
        cur = con.cursor()
        cur.execute('SELECT * FROM student')
        lst_code = cur.fetchall()
        for i in range(0, len(lst_code)):       #har tuple
            for j in range(0, len(lst_code[i])):        #har item tuple
                s.cell(row=i+2, column=j+1).value = lst_code[i][j]
        w.save('c:\\haniye\\student.xlsx')
    except Exception as x:
        print(x)


def back_up_sale():
    w = openpyxl.Workbook()
    s = w.active
    s.title = 'sale'
    s['a1'] = 'name'
    s['b1'] = 'last name'
    s['c1'] = 'code'
    s['d1'] = 'bday'
    s['e1'] = 'phone num'
    s['f1'] = 'madrak'
    s['g1'] = 'shift'
    s['h1'] = 'percent'
    s['i1'] = 'bimeh'
    s['j1'] = 'sabeqe'
    s['k1'] = 'checks'
    s['l1'] = 'zamen'
    try:                                                    #etesal be databse(mysql)

        con = sql.connect(host='localhost', user='root', password='P@ssw0rd', database="amozeshgah")
        cur = con.cursor()
        cur.execute('SELECT * FROM sale')
        lst_code = cur.fetchall()
        for i in range(0, len(lst_code)):                #har tuple
            for j in range(0, len(lst_code[i])):        #har item tuple
                s.cell(row=i+2, column=j+1).value = lst_code[i][j]
        w.save('c:\\haniye\\sale.xlsx')
    except Exception as x:
        print(x)


def back_up_fanni():
    w = openpyxl.Workbook()
    s = w.active
    s.title = 'sale'
    s['a1'] = 'name'
    s['b1'] = 'last name'
    s['c1'] = 'code'
    s['d1'] = 'bday'
    s['e1'] = 'phone num'
    s['f1'] = 'madrak'
    s['g1'] = 'shift'
    s['h1'] = 'skill'
    s['i1'] = 'bimeh'
    s['j1'] = 'sabeqe'
    s['k1'] = 'checks'
    s['l1'] = 'zamen'
    try:                                                    #etesal be databse(mysql)

        con = sql.connect(host='localhost', user='root', password='P@ssw0rd', database="amozeshgah")
        cur = con.cursor()
        cur.execute('SELECT * FROM fanni')
        lst_code = cur.fetchall()
        for i in range(0, len(lst_code)):                #har tuple
            for j in range(0, len(lst_code[i])):        #har item tuple
                s.cell(row=i+2, column=j+1).value = lst_code[i][j]
        w.save('c:\\haniye\\fanni.xlsx')
    except Exception as x:
        print(x)


def back_up_classes():
    w = openpyxl.Workbook()
    s = w.active
    s.title = "classes"
    s['a1'] = 'calss number'
    s['b1'] = 'zarfiat'
    s['c1'] = 'tedad system'
    s['d1'] = 'tabaqe'

    try:                                                    #etesal be databse(mysql)

        con = sql.connect(host='localhost', user='root', password='P@ssw0rd', database="amozeshgah")
        cur = con.cursor()
        cur.execute('SELECT * FROM classes')
        lst_code = cur.fetchall()
        for i in range(0, len(lst_code)):       #har tuple
            for j in range(0, len(lst_code[i])):        #har item tuple
                s.cell(row=i+2, column=j+1).value = lst_code[i][j]
        w.save('c:\\haniye\\classes.xlsx')
    except Exception as x:
        print(x)


def back_up_lessons():
    w = openpyxl.Workbook()
    s = w.active
    s.title = "lessons"
    s['a1'] = 'code'
    s['b1'] = 'lesson name'
    s['c1'] = 'saat'
    s['d1'] = 'type'

    try:                                                    #etesal be databse(mysql)
        con = sql.connect(host='localhost', user='root', password='P@ssw0rd', database="amozeshgah")
        cur = con.cursor()
        cur.execute('SELECT * FROM lessons')
        lst_code = cur.fetchall()
        for i in range(0, len(lst_code)):       #har tuple
            for j in range(0, len(lst_code[i])):        #har item tuple
                s.cell(row=i+2, column=j+1).value = lst_code[i][j]
        w.save('c:\\haniye\\lessons.xlsx')
    except Exception as x:
        print(x)

#back_up_teacher()
#back_up_student()
#back_up_fanni()
#back_up_sale()
#back_up_lessons()
#back_up_classes()


if __name__ == '__main__':
    print('its module not the main program')



